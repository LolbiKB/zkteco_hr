"""
Spreadsheet-based bulk schedule import.

Accepts a CSV (or xlsx) in the canonical import format:

  employee_id  | email | am_from | am_to | pm_from | pm_to | days_off

Use the Haiku prompt in docs/SCHEDULE_IMPORT_PROMPT.md to normalise any raw
spreadsheet into this format before importing.

Exposed API: parse_schedule_upload(file_b64, filename)
"""

from __future__ import annotations

import base64
import csv
import io
import re
from dataclasses import dataclass, field
from typing import Literal

import frappe

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_WEEKDAY_IDX: dict[str, int] = {d: i for i, d in enumerate(WEEKDAYS)}

_DAY_ALIASES: dict[str, str] = {
    "mon": "Monday", "monday": "Monday",
    "tue": "Tuesday", "tues": "Tuesday", "tuesday": "Tuesday",
    "wed": "Wednesday", "wednesday": "Wednesday",
    "thu": "Thursday", "thur": "Thursday", "thurs": "Thursday", "thursday": "Thursday",
    "fri": "Friday", "friday": "Friday",
    "sat": "Saturday", "saturday": "Saturday",
    "sun": "Sunday", "sunday": "Sunday",
}

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})$")
_ID_PATTERN = re.compile(r"^[A-Za-z]{1,4}-\d{2,}", re.IGNORECASE)
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_DATE_LIKE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_DURATION_RE = re.compile(r"^\d+h$", re.IGNORECASE)

ScheduleShape = Literal["full_day", "am_only", "pm_only", "continuous", "invalid"]
IssueSeverity = Literal["error", "warning", "info"]

# Machine-readable codes — surfaced in UI + feedback export for the AI normaliser.
ISSUE_CODES = {
    "MISSING_EMPLOYEE_ID": "error",
    "INVALID_EMPLOYEE_ID": "error",
    "EMPLOYEE_NOT_FOUND": "error",
    "INVALID_TIME_FORMAT": "error",
    "MISSING_SHIFT_TIMES": "error",
    "END_BEFORE_START": "error",
    "NO_WORKING_DAYS": "error",
    "GARBAGE_ROW": "error",
    "MIDNIGHT_AS_NOON": "warning",
    "INVALID_EMAIL": "warning",
    "INVALID_DAYS_OFF_TOKEN": "warning",
    "DUPLICATE_EMPLOYEE_ID": "warning",
    "SHORT_LUNCH_GAP": "warning",
    "PM_ONLY": "info",
    "CONTINUOUS_SHIFT": "info",
    "AM_ONLY": "info",
}

AI_SUGGESTIONS: dict[str, str] = {
    "MISSING_EMPLOYEE_ID": "Set employee_id to the badge number (e.g. DI-0159). Do not leave blank.",
    "INVALID_EMPLOYEE_ID": "employee_id must look like DI-1234 or DIS-1234 — not a date, time, or email.",
    "EMPLOYEE_NOT_FOUND": "Verify the badge exists in Frappe Employee (employee_number or attendance_device_id).",
    "INVALID_TIME_FORMAT": "Use HH:MM 24-hour only (07:30, 12:00, 17:00). No am/pm suffixes.",
    "MISSING_SHIFT_TIMES": "Provide either AM times, PM times, or a continuous span (am_from + pm_to with off middle).",
    "END_BEFORE_START": "End time must be after start time on the same day.",
    "NO_WORKING_DAYS": "days_off lists every weekday — employee would never work. Fix days_off or shift times.",
    "GARBAGE_ROW": "Row failed normalisation — remove or re-run Haiku on the source row.",
    "MIDNIGHT_AS_NOON": "00:00 is likely noon — use 12:00 for lunch end when morning starts 06:00–11:00.",
    "INVALID_EMAIL": "email must be a single address (name@domain.tld). Fix commas or stray text.",
    "INVALID_DAYS_OFF_TOKEN": "days_off uses pipe-separated full weekday names; append (am) for morning-only off.",
    "DUPLICATE_EMPLOYEE_ID": "Same employee_id appears on multiple rows — dedupe before import.",
    "SHORT_LUNCH_GAP": "Lunch gap under 15 minutes — confirm am_to and pm_from are correct.",
    "PM_ONLY": "Afternoon-only employee — am_from/am_to should both be off.",
    "CONTINUOUS_SHIFT": "Long span without lunch — am_to/pm_from are off; using am_from→pm_to.",
    "AM_ONLY": "Morning-only employee — pm_from/pm_to should both be off.",
}


@dataclass
class ImportIssue:
    code: str
    severity: IssueSeverity
    message: str
    field: str | None = None
    suggestion: str | None = None

    def to_dict(self) -> dict:
        return {
            "code": self.code,
            "severity": self.severity,
            "message": self.message,
            "field": self.field,
            "suggestion": self.suggestion or AI_SUGGESTIONS.get(self.code),
        }


@dataclass
class ParsedScheduleRow:
    row_number: int
    id_card: str
    email: str
    employee: str | None = None
    employee_name: str | None = None
    matched: bool = False
    am_from: str | None = None
    am_to: str | None = None
    pm_from: str | None = None
    pm_to: str | None = None
    day_off: dict = field(default_factory=lambda: {"full_off": [], "afternoon_off": []})
    week_pattern: dict | None = None
    schedule_shape: ScheduleShape = "invalid"
    issues: list[ImportIssue] = field(default_factory=list)
    importable: bool = False

    @property
    def warnings(self) -> list[str]:
        """Back-compat string warnings for older clients."""
        return [i.message for i in self.issues if i.severity in ("error", "warning")]

    def to_dict(self) -> dict:
        return {
            "row_number": self.row_number,
            "id_card": self.id_card,
            "email": self.email,
            "employee": self.employee,
            "employee_name": self.employee_name,
            "matched": self.matched,
            "am_from": self.am_from,
            "am_to": self.am_to,
            "pm_from": self.pm_from,
            "pm_to": self.pm_to,
            "day_off": self.day_off,
            "week_pattern": self.week_pattern,
            "schedule_shape": self.schedule_shape,
            "issues": [i.to_dict() for i in self.issues],
            "importable": self.importable,
            "warnings": self.warnings,
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalize_weekday(s: str) -> str | None:
    return _DAY_ALIASES.get(s.strip().lower())


def _is_blank(value) -> bool:
    return not value or str(value).strip().lower() in ("", "off", "none", "n/a", "-")


def _cell(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("nan", "none") else s


def _parse_time(value, *, field: str, issues: list[ImportIssue]) -> str | None:
    text = _cell(value)
    if not text or _is_blank(text):
        return None
    m = _TIME_RE.match(text)
    if not m:
        issues.append(ImportIssue(
            code="INVALID_TIME_FORMAT",
            severity="error",
            message=f"{field} has invalid time {text!r} — expected HH:MM 24h.",
            field=field,
        ))
        return None
    hours, minutes = int(m.group(1)), int(m.group(2))
    if hours > 23 or minutes > 59:
        issues.append(ImportIssue(
            code="INVALID_TIME_FORMAT",
            severity="error",
            message=f"{field} time out of range: {text!r}.",
            field=field,
        ))
        return None
    normalized = f"{hours:02d}:{minutes:02d}"
    if normalized == "00:00" and field in ("am_to", "pm_from"):
        issues.append(ImportIssue(
            code="MIDNIGHT_AS_NOON",
            severity="warning",
            message=f"{field} is 00:00 — often a noon misread; consider 12:00.",
            field=field,
        ))
    return normalized


def _time_to_minutes(value: str | None) -> int | None:
    if not value:
        return None
    parts = value.split(":")
    if len(parts) != 2:
        return None
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        return None


def _parse_days_off(raw: str, issues: list[ImportIssue]) -> tuple[list[str], list[str], list[str]]:
    full_off: list[str] = []
    am_only: list[str] = []
    invalid_tokens: list[str] = []

    if not raw or not raw.strip():
        return full_off, am_only, invalid_tokens

    for token in re.split(r"[|;,]", raw):
        token = token.strip()
        if not token:
            continue
        is_am = bool(re.search(r"\(am\)", token, re.IGNORECASE))
        day_name = re.sub(r"\(am\)", "", token, flags=re.IGNORECASE).strip()
        weekday = _normalize_weekday(day_name)
        if not weekday:
            invalid_tokens.append(token)
            continue
        if is_am:
            am_only.append(weekday)
        else:
            full_off.append(weekday)

    if invalid_tokens:
        issues.append(ImportIssue(
            code="INVALID_DAYS_OFF_TOKEN",
            severity="warning",
            message=f"Unrecognised days_off token(s): {', '.join(invalid_tokens)}.",
            field="days_off",
        ))

    return (
        sorted(full_off, key=lambda d: _WEEKDAY_IDX[d]),
        sorted(am_only, key=lambda d: _WEEKDAY_IDX[d]),
        invalid_tokens,
    )


def _looks_like_garbage_id(id_card: str, email: str) -> bool:
    if not id_card:
        return False
    if _TIME_RE.match(id_card) or _DATE_LIKE_RE.match(id_card) or _DURATION_RE.match(id_card):
        return True
    if id_card.lower() in ("no", "4h"):
        return True
    if "@" in id_card and not _ID_PATTERN.match(id_card):
        return True
    return False


def _validate_email(email: str, issues: list[ImportIssue]) -> None:
    if not email:
        return
    cleaned = email.strip().strip('"')
    if cleaned.lower() in ("no", "n/a"):
        issues.append(ImportIssue(
            code="INVALID_EMAIL",
            severity="warning",
            message=f"email looks invalid: {email!r}.",
            field="email",
        ))
        return
    if "," in cleaned and "@" in cleaned:
        issues.append(ImportIssue(
            code="INVALID_EMAIL",
            severity="warning",
            message=f"email contains comma — likely CSV corruption: {email!r}.",
            field="email",
        ))
        return
    if "@" in cleaned and not _EMAIL_RE.match(cleaned):
        issues.append(ImportIssue(
            code="INVALID_EMAIL",
            severity="warning",
            message=f"email format suspicious: {email!r}.",
            field="email",
        ))


def _detect_schedule_shape(
    am_from: str | None,
    am_to: str | None,
    pm_from: str | None,
    pm_to: str | None,
    issues: list[ImportIssue],
) -> tuple[ScheduleShape, str | None, str | None, str | None, str | None]:
    """Normalise raw cells into a schedule shape and resolved time quadruple."""
    am_start = am_from if not _is_blank(am_from) else None
    am_end = am_to if not _is_blank(am_to) else None
    pm_start = pm_from if not _is_blank(pm_from) else None
    pm_end = pm_to if not _is_blank(pm_to) else None

    # Continuous: 06:00, off, off, 18:00
    if am_start and not am_end and not pm_start and pm_end:
        issues.append(ImportIssue(
            code="CONTINUOUS_SHIFT",
            severity="info",
            message=f"Continuous shift {am_start}–{pm_end} (no lunch break).",
        ))
        return "continuous", am_start, None, None, pm_end

    # PM-only: off, off, 13:00, 17:00
    if not am_start and not am_end and pm_start and pm_end:
        issues.append(ImportIssue(
            code="PM_ONLY",
            severity="info",
            message=f"Afternoon-only shift {pm_start}–{pm_end}.",
        ))
        return "pm_only", None, None, pm_start, pm_end

    # AM-only: 07:00, 11:30, off, off
    if am_start and am_end and not pm_start and not pm_end:
        issues.append(ImportIssue(
            code="AM_ONLY",
            severity="info",
            message=f"Morning-only shift {am_start}–{am_end}.",
        ))
        return "am_only", am_start, am_end, None, None

    # Full day with lunch split
    if am_start and am_end and pm_start and pm_end:
        return "full_day", am_start, am_end, pm_start, pm_end

    # Partial PM (am set but pm incomplete)
    if am_start and am_end and not pm_start and not pm_end:
        return "am_only", am_start, am_end, None, None

    if not am_start and not am_end and (pm_start or pm_end):
        issues.append(ImportIssue(
            code="MISSING_SHIFT_TIMES",
            severity="error",
            message="PM shift incomplete — need both pm_from and pm_to.",
            field="pm_from",
        ))
        return "invalid", am_start, am_end, pm_start, pm_end

    issues.append(ImportIssue(
        code="MISSING_SHIFT_TIMES",
        severity="error",
        message="No complete shift window — check am_from/am_to and pm_from/pm_to.",
    ))
    return "invalid", am_start, am_end, pm_start, pm_end


def _validate_time_order(
    shape: ScheduleShape,
    am_from: str | None,
    am_to: str | None,
    pm_from: str | None,
    pm_to: str | None,
    issues: list[ImportIssue],
) -> None:
    if shape == "full_day" and am_from and am_to and pm_from and pm_to:
        am_end_m = _time_to_minutes(am_to)
        pm_start_m = _time_to_minutes(pm_from)
        if am_end_m is not None and pm_start_m is not None:
            gap = pm_start_m - am_end_m
            if gap <= 0:
                issues.append(ImportIssue(
                    code="END_BEFORE_START",
                    severity="error",
                    message=f"Lunch gap invalid: am_to {am_to} is not before pm_from {pm_from}.",
                    field="pm_from",
                ))
            elif gap < 15:
                issues.append(ImportIssue(
                    code="SHORT_LUNCH_GAP",
                    severity="warning",
                    message=f"Lunch gap only {gap} min ({am_to}–{pm_from}).",
                    field="pm_from",
                ))

    spans = {
        "am_only": (am_from, am_to),
        "pm_only": (pm_from, pm_to),
        "continuous": (am_from, pm_to),
        "full_day": (am_from, pm_to),
    }
    start, end = spans.get(shape, (None, None))
    if start and end:
        start_m, end_m = _time_to_minutes(start), _time_to_minutes(end)
        if start_m is not None and end_m is not None and end_m <= start_m:
            issues.append(ImportIssue(
                code="END_BEFORE_START",
                severity="error",
                message=f"Shift end {end} is not after start {start}.",
                field="pm_to" if shape in ("continuous", "full_day") else "am_to",
            ))


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------


def _read_xlsx(file_bytes: bytes) -> list[list]:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return [[_cell(cell.value) for cell in row] for row in ws.iter_rows()]


def _read_csv(file_bytes: bytes) -> list[list]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _find_data_start(rows: list[list]) -> int:
    for i, row in enumerate(rows):
        if not row:
            continue
        first = _cell(row[0]).lower()
        if first in ("employee_id", "employee", "id", "badge"):
            return i + 1
        if _ID_PATTERN.match(_cell(row[0])):
            return i
        # Rows with empty id but valid email+times are handled later
    return len(rows)


# ---------------------------------------------------------------------------
# Employee lookup
# ---------------------------------------------------------------------------


def _lookup_employee(id_card: str, email: str = "") -> tuple[str | None, str | None]:
    for fieldname in ("employee_number", "attendance_device_id"):
        result = frappe.db.get_value(
            "Employee",
            {fieldname: id_card, "status": "Active"},
            ["name", "employee_name"],
            as_dict=True,
        )
        if result:
            return result["name"], result["employee_name"]

    if email:
        cleaned = email.strip().strip('"')
        for email_field in ("company_email", "personal_email"):
            result = frappe.db.get_value(
                "Employee",
                {email_field: cleaned, "status": "Active"},
                ["name", "employee_name"],
                as_dict=True,
            )
            if result:
                return result["name"], result["employee_name"]

    return None, None


# ---------------------------------------------------------------------------
# WeekPattern builder
# ---------------------------------------------------------------------------


def _build_week_pattern(
    shape: ScheduleShape,
    am_from: str | None,
    am_to: str | None,
    pm_from: str | None,
    pm_to: str | None,
    full_off: list[str],
    am_only: list[str],
) -> dict | None:
    if shape == "invalid":
        return None

    full_off_set = set(full_off)
    am_only_set = set(am_only)

    days = []
    working_count = 0
    for weekday in WEEKDAYS:
        if weekday in full_off_set:
            days.append({"weekday": weekday, "works": False})
            continue

        working_count += 1

        if shape == "pm_only":
            days.append({
                "weekday": weekday, "works": True,
                "start_time": pm_from, "end_time": pm_to,
                "lunch_start": None, "lunch_end": None,
                "grace_minutes": 10,
            })
        elif shape == "continuous":
            days.append({
                "weekday": weekday, "works": True,
                "start_time": am_from, "end_time": pm_to,
                "lunch_start": None, "lunch_end": None,
                "grace_minutes": 10,
            })
        elif shape == "am_only" or weekday in am_only_set:
            days.append({
                "weekday": weekday, "works": True,
                "start_time": am_from, "end_time": am_to,
                "lunch_start": None, "lunch_end": None,
                "grace_minutes": 10,
            })
        else:
            days.append({
                "weekday": weekday, "works": True,
                "start_time": am_from, "end_time": pm_to,
                "lunch_start": am_to, "lunch_end": pm_from,
                "grace_minutes": 10,
            })

    if working_count == 0:
        return None

    return {"frequency": "Every Week", "days": days}


def _parse_row(raw: list, row_number: int) -> ParsedScheduleRow:
    while len(raw) < 7:
        raw.append("")

    id_card = _cell(raw[0])
    email = _cell(raw[1]).strip('"')
    issues: list[ImportIssue] = []

    row = ParsedScheduleRow(row_number=row_number, id_card=id_card, email=email, issues=issues)

    # Skip completely empty rows
    if not any(_cell(c) for c in raw[:7]):
        return row

    if _looks_like_garbage_id(id_card, email):
        issues.append(ImportIssue(
            code="GARBAGE_ROW",
            severity="error",
            message=f"Row {row_number} looks like failed normalisation (id={id_card!r}).",
            field="employee_id",
        ))

    if not id_card:
        pass  # handled after employee lookup
    elif not _ID_PATTERN.match(id_card):
        issues.append(ImportIssue(
            code="INVALID_EMPLOYEE_ID",
            severity="error",
            message=f"employee_id {id_card!r} does not match expected badge format (e.g. DI-0159).",
            field="employee_id",
        ))

    _validate_email(email, issues)

    am_from = _parse_time(raw[2], field="am_from", issues=issues)
    am_to = _parse_time(raw[3], field="am_to", issues=issues)
    pm_from = _parse_time(raw[4], field="pm_from", issues=issues) if not _is_blank(raw[4]) else None
    pm_to = _parse_time(raw[5], field="pm_to", issues=issues) if not _is_blank(raw[5]) else None

    full_off, am_only, _invalid = _parse_days_off(_cell(raw[6]), issues)

    shape, am_from, am_to, pm_from, pm_to = _detect_schedule_shape(
        am_from, am_to, pm_from, pm_to, issues
    )
    _validate_time_order(shape, am_from, am_to, pm_from, pm_to, issues)

    row.am_from = am_from
    row.am_to = am_to
    row.pm_from = pm_from
    row.pm_to = pm_to
    row.schedule_shape = shape
    row.day_off = {"full_off": full_off, "afternoon_off": am_only}

    week_pattern = _build_week_pattern(shape, am_from, am_to, pm_from, pm_to, full_off, am_only)
    if week_pattern:
        working = sum(1 for d in week_pattern["days"] if d.get("works"))
        if working == 0:
            issues.append(ImportIssue(
                code="NO_WORKING_DAYS",
                severity="error",
                message="All weekdays marked off — no shift to import.",
                field="days_off",
            ))
            week_pattern = None
    row.week_pattern = week_pattern

    employee, employee_name = (None, None)
    if id_card:
        employee, employee_name = _lookup_employee(id_card, email)
    elif email and "@" in email:
        employee, employee_name = _lookup_employee("", email)

    if not id_card:
        if employee:
            issues.append(ImportIssue(
                code="MISSING_EMPLOYEE_ID",
                severity="warning",
                message=f"Row {row_number} matched by email only — add employee_id to CSV.",
                field="employee_id",
            ))
        else:
            issues.append(ImportIssue(
                code="MISSING_EMPLOYEE_ID",
                severity="error",
                message=f"Row {row_number} has no employee_id.",
                field="employee_id",
            ))

    row.employee = employee
    row.employee_name = employee_name
    row.matched = bool(employee)

    if id_card and not employee:
        issues.append(ImportIssue(
            code="EMPLOYEE_NOT_FOUND",
            severity="error",
            message=f"No active employee found for {id_card!r}.",
            field="employee_id",
        ))
    elif not id_card and not employee and email:
        issues.append(ImportIssue(
            code="EMPLOYEE_NOT_FOUND",
            severity="error",
            message=f"No active employee found for email {email!r}.",
            field="email",
        ))

    has_error = any(i.severity == "error" for i in issues)
    row.importable = bool(row.matched and week_pattern and not has_error)
    return row


def _apply_duplicate_warnings(rows: list[ParsedScheduleRow]) -> None:
    seen: dict[str, list[int]] = {}
    for row in rows:
        key = row.id_card.strip().upper()
        if not key:
            continue
        seen.setdefault(key, []).append(row.row_number)

    for id_card, row_numbers in seen.items():
        if len(row_numbers) < 2:
            continue
        for row in rows:
            if row.id_card.strip().upper() != id_card:
                continue
            row.issues.append(ImportIssue(
                code="DUPLICATE_EMPLOYEE_ID",
                severity="warning",
                message=f"Duplicate employee_id {id_card} on rows {', '.join(map(str, row_numbers))}.",
                field="employee_id",
            ))
            if any(i.severity == "error" for i in row.issues):
                row.importable = False


def _build_summary(rows: list[ParsedScheduleRow]) -> dict:
    by_code: dict[str, int] = {}
    for row in rows:
        for issue in row.issues:
            by_code[issue.code] = by_code.get(issue.code, 0) + 1

    return {
        "total_rows": len(rows),
        "importable": sum(1 for r in rows if r.importable),
        "matched": sum(1 for r in rows if r.matched),
        "unmatched": sum(1 for r in rows if r.id_card and not r.matched),
        "errors": sum(1 for r in rows if any(i.severity == "error" for i in r.issues)),
        "warnings": sum(1 for r in rows if any(i.severity == "warning" for i in r.issues)),
        "garbage_rows": sum(1 for r in rows if any(i.code == "GARBAGE_ROW" for i in r.issues)),
        "by_code": by_code,
    }


def _build_feedback_rows(rows: list[ParsedScheduleRow]) -> list[dict]:
    feedback: list[dict] = []
    for row in rows:
        for issue in row.issues:
            feedback.append({
                "row_number": row.row_number,
                "employee_id": row.id_card,
                "email": row.email,
                "field": issue.field or "",
                "code": issue.code,
                "severity": issue.severity,
                "message": issue.message,
                "suggestion": issue.suggestion or AI_SUGGESTIONS.get(issue.code, ""),
            })
    return feedback


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


@frappe.whitelist()
def parse_schedule_upload(file_b64: str, filename: str = "upload.csv") -> dict:
    """
    Parse a base64-encoded canonical schedule CSV/xlsx and return a preview.

    Expected columns (header row optional):
      employee_id, email, am_from, am_to, pm_from, pm_to, days_off
    """
    try:
        file_bytes = base64.b64decode(file_b64)
    except Exception as exc:
        frappe.throw(f"Invalid file data: {exc}")

    ext = (filename or "").lower().rsplit(".", 1)[-1]
    try:
        raw_rows = _read_csv(file_bytes) if ext == "csv" else _read_xlsx(file_bytes)
    except Exception as exc:
        frappe.throw(f"Failed to read file: {exc}")

    data_start = _find_data_start(raw_rows)
    if data_start >= len(raw_rows):
        frappe.throw(
            "No employee rows found. "
            "First column must be an employee ID (e.g. DI-0159). "
            "Use the Haiku prompt to normalise your spreadsheet first."
        )

    parsed_rows: list[ParsedScheduleRow] = []
    for file_idx, raw in enumerate(raw_rows[data_start:], start=data_start + 1):
        row = _parse_row(raw, file_idx)
        if not row.id_card and not row.email and not row.week_pattern:
            if not row.issues:
                continue
        parsed_rows.append(row)

    _apply_duplicate_warnings(parsed_rows)

    # Recompute importable after duplicate warnings
    for row in parsed_rows:
        has_error = any(i.severity == "error" for i in row.issues)
        row.importable = bool(row.matched and row.week_pattern and not has_error)

    summary = _build_summary(parsed_rows)
    feedback_rows = _build_feedback_rows(parsed_rows)

    return {
        "rows": [r.to_dict() for r in parsed_rows],
        "summary": summary,
        "feedback_rows": feedback_rows,
        "normalized_by": "rules",
    }
